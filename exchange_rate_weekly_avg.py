import requests
import pandas as pd
import smtplib
from email.message import EmailMessage
from pathlib import Path
from datetime import date, timedelta
from openpyxl import load_workbook


base_url = "https://eservices.mas.gov.sg/apimg-gw/server/monthly_statistical_bulletin_non610ora/exchange_rates_average_for_period_weekly/views/exchange_rates_average_for_period_weekly"
headers = {
    "accept": "application/json; charset=UTF-8",
    "KeyId": "dd277b45-55c5-4c1d-8d5e-2937466f8bf6"
    }
today = date.today()

for i in range(8):
    check_date = today - timedelta(days=i)
    end_of_week_param = check_date.strftime("%Y-%m-%d")

    params = {
        "end_of_week": end_of_week_param
    }

    response = requests.get(base_url, headers=headers, params=params)

    if response.status_code != 200:
        continue

    try:
        data = response.json()
    except:
        continue

    elements = data.get("elements", [])

    if elements:
        print("Latest available date found:", end_of_week_param)
        df = pd.DataFrame(elements)
        selected_columns = [
    "end_of_week",
    "aud_sgd", 
    "chf_sgd", 
    "cny_sgd_100",
    "eur_sgd",
    "gbp_sgd",
    "jpy_sgd_100",
    "usd_sgd",
    ]

        df_selected = df[selected_columns]

        # Convert all rate columns to numeric first
        rate_columns = selected_columns[1:]
        df_selected[rate_columns] = df_selected[rate_columns].apply(pd.to_numeric, errors="coerce")

        # Convert per 100 currency rates to per 1 unit
        df_selected["cny_sgd_100"] = df_selected["cny_sgd_100"] * 0.01
        df_selected["jpy_sgd_100"] = df_selected["jpy_sgd_100"] * 0.01
      

        print(df_selected)
        break
else:
    raise Exception("No exchange rate data found in the last 7 days.")


df_selected = df_selected.rename(columns={
    "end_of_week": "Date",
    "eur_sgd": "EUR",
    "gbp_sgd": "GBP",
    "usd_sgd": "USD",
    "aud_sgd": "AUD",
    "cny_sgd_100": "CNY",
    "jpy_sgd_100": "JPY",
    "chf_sgd": "CHF"
})



# Insert empty SEK column between JPY and USD
df_selected.insert(
    loc=df_selected.columns.get_loc("USD"),
    column="SEK",
    value=""
)

date = today + timedelta(days=3)
enddate = today + timedelta(days=9)
date = date.strftime("%d %b %Y")
enddate = enddate.strftime("%d %b %Y")
week = " ("+ date + " - " + enddate + ")"
# Export to Excel
output_file = r"W:\ACCTS\Exchange Rate\Year 2026\MAS weekly rates{week}.xlsx".format(week=week)

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df_selected.to_excel(writer, sheet_name="Exchange_Rates", index=False)
   
print("Excel file created successfully.")

#--------control / (cmt off)---------------------------------------------- Email sending with attachment --------------------------------------------------------------------------------------------

# file_path = Path(output_file)

# # Check file exists before sending
# if not file_path.exists():
#     raise FileNotFoundError("Excel file was not generated. Email not sent.")

# recipients = [
#     "koksharon@faxolif.com",
#     "kudsiah@faxolif.com",
#     "khinmlk@faxolif.com",
#     "sktan@faxolif.com",
#     "yjchin@faxolif.com",
#     "emilykhaing@faxolif.com",
#     "sharonszeto@faxolif.com",
#     "jqchu@faxolif.com",
#     "leannsim@faxolif.com",
#     "maslinda@faxolif.com",
#     "ykcheong@faxolif.com",
#     "mcyong@faxolif.com",
#     "vadiveilan@faxolif.com"

# ]

# msg = EmailMessage()
# msg["Subject"] = "MAS Weekly Exchange Rate{week}".format(week=week)
# msg["From"] = "emilykhaing@faxolif.com"
# msg["To"] = ", ".join(recipients)

# #koksharon@faxolif.com, kudsiah@faxolif.com, khinmlk@faxolif.com, sktan@faxolif.com, 
# #sharonszeto@faxolif.com, carmenlow@faxolif.com, ykcheong@faxolif.com, 

# msg.set_content("""
# Dear All,

# Please find attached the latest automated MAS exchange rate report.
# (All exchange rates are based against SGD 1.)

# Kindly let me know if you have any questions or feedback.

# Best regards,
# Emily
# """)

# # Attach Excel file
# with open(file_path, "rb") as f:
#     msg.add_attachment(
#         f.read(),
#         maintype="application",
#         subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         filename=file_path.name
#     )

# # SMTP sending
# with smtplib.SMTP("mail2.trafomaterials.sg", 587) as server:
#     server.starttls()
#     server.login("emilykhaing@faxolif.com", "keL6?m!Yna09")
#     server.send_message(msg)

# print("Email sent successfully.")


# Take latest API row to update SAP upload page ---------------------------------------------------------------------------------------------------------------
latest = df_selected.iloc[0]

api_date = latest["Date"]

# Since MAS rate is Friday, apply from next Monday
start_date = pd.to_datetime(api_date) + pd.Timedelta(days=3)    
# Monday = 0, Friday = 4, so 2026-05-15 + 3 days = 2026-05-18

# Duplicate for 7 days
df_7days = pd.DataFrame({
    "Date": pd.date_range(start=start_date, periods=7, freq="D")
})

# Copy all currency values to 7 rows
for col in df_selected.columns:
    if col != "Date":
        df_7days[col] = latest[col]

#print(df_7days)

datemonth = today + timedelta(days=3)  # Next Monday
update_sap_path = r"W:\ACCTS\Exchange Rate\Year 2026\SAP Exchange Rates 2026.xlsx"
wb = load_workbook(update_sap_path)

# Sheet name based on month, example: May

monthname = datemonth.strftime("%b").upper()  # MAY
sheet_name = monthname
ws = wb[sheet_name]
#print(ws)

wb = load_workbook(update_sap_path)

for _, row in df_7days.iterrows():
    date_value = pd.to_datetime(row["Date"])
    day_number = date_value.day

    # Your Excel sheets are named JUN, JUL, AUG
    sheet_name = date_value.strftime("%b").upper()

    if sheet_name not in wb.sheetnames:
        print(f"Sheet not found: {sheet_name}")
        continue

    ws = wb[sheet_name]

    # Build header mapping
    header_row = 1
    header_map = {}

    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=header_row, column=col).value
        if header_value:
            header_map[str(header_value).strip()] = col

    # Find matching day number in column A
    target_row = None

    for excel_row in range(2, ws.max_row + 1):
        if ws.cell(row=excel_row, column=1).value == day_number:
            target_row = excel_row
            break

    if target_row is None:
        print(f"Day {day_number} not found in sheet {sheet_name}")
        continue

    # Update currency columns
    for currency in df_7days.columns:
        if currency == "Date":
            continue

        if currency in header_map:
            ws.cell(
                row=target_row,
                column=header_map[currency]
            ).value = row[currency]

wb.save(update_sap_path)

print("Update to SAP File updated successfully.")


#---------------------------------------------------------- to upload to SAP Shared folder ---------------------------------------------------------------------------------------------------------------

#I dont need SEK with NULL values if i use DTW, so I will not add SEK to the dataframe.
date_value = pd.to_datetime(df_selected["Date"].iloc[0], errors="coerce") + pd.Timedelta(days=3)

currency_df = pd.DataFrame({
    "Currency": [df_selected.columns[i] for i in range(1, len(df_selected.columns))],
    "date": date_value.strftime("%m/%d/%Y"),
    "Rate": [df_selected.iloc[0, i] for i in range(1, len(df_selected.columns))]
})
#print(currency_df)

# Duplicate for 7 days
date_value = pd.to_datetime(df_selected["Date"].iloc[0], errors="coerce") + pd.Timedelta(days=3)
currency_toupload = []

for d in pd.date_range(start=date_value, periods=7, freq="D"):
        temp_df = pd.DataFrame({
            "Currency": [df_selected.columns[j] for j in range(1, len(df_selected.columns))],
            "date": d.strftime("%m/%d/%Y"),
            "Rate": [df_selected.iloc[0, j] for j in range(1, len(df_selected.columns))]
        })
        currency_toupload.append(temp_df)

currency_toupload = pd.concat(currency_toupload, ignore_index=True)        
print(currency_toupload)


    # Duplicate the first data row
first_row = currency_toupload.iloc[[0]]

currency_toupload_final = pd.concat(
    [first_row, currency_toupload],
    ignore_index=True
)
#txtfile_path = r"C:\Users\Emily\Desktop\Query\ExchangeRate_UplodtoSAP.txt"
#currency_toupload_final.to_csv(txtfile_path, index=False, header=True, sep="\t")

currency_toupload_final = pd.concat(
    [first_row, currency_toupload],
    ignore_index=True
)
txtfile_path2 = r"Y:\ExchangeRateUpdate.txt"

currency_toupload_final.to_csv(txtfile_path2, index=False, header=True, sep="\t")

print("File has been saved to SAP Shared folder.")